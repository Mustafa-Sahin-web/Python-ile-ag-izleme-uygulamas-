"""
Excel Exporter Module
Tarama sonuçlarını Excel'e aktaran modül
"""
from typing import List, Dict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime


class ExcelExporter:
    """Excel dosyasına veri aktarımı yapan sınıf"""
    
    def __init__(self):
        """ExcelExporter başlatıcı"""
        self.workbook = None
        self.worksheet = None
    
    def create_workbook(self):
        """Yeni bir Excel çalışma kitabı oluşturur"""
        self.workbook = Workbook()
        self.worksheet = self.workbook.active
        self.worksheet.title = "Ağ Tarama Sonuçları"
    
    def set_headers(self):
        """Excel tablosuna başlıkları ekler"""
        headers = ['IP Adresi', 'Hostname', 'Durum', 'Açık Portlar']
        
        # Başlık hücrelerini ayarla
        for col, header in enumerate(headers, start=1):
            cell = self.worksheet.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Sütun genişliklerini ayarla
        self.worksheet.column_dimensions['A'].width = 15
        self.worksheet.column_dimensions['B'].width = 30
        self.worksheet.column_dimensions['C'].width = 12
        self.worksheet.column_dimensions['D'].width = 30
    
    def add_data(self, results: List[Dict]):
        """
        Tarama sonuçlarını Excel'e ekler
        
        Args:
            results: Tarama sonuçları listesi
        """
        for row_idx, result in enumerate(results, start=2):
            self.worksheet.cell(row=row_idx, column=1, value=result['ip'])
            self.worksheet.cell(row=row_idx, column=2, value=result['hostname'])
            self.worksheet.cell(row=row_idx, column=3, value=result['status'])
            self.worksheet.cell(row=row_idx, column=4, value=result['open_ports'])
            
            # Hücreleri ortala
            for col in range(1, 5):
                cell = self.worksheet.cell(row=row_idx, column=col)
                cell.alignment = Alignment(horizontal="left", vertical="center")
    
    def add_summary(self, ip_block: str, total_hosts: int):
        """
        Özet bilgileri ekler
        
        Args:
            ip_block: Taranan IP bloğu
            total_hosts: Bulunan toplam host sayısı
        """
        last_row = self.worksheet.max_row + 2
        
        self.worksheet.cell(row=last_row, column=1, value="Özet Bilgiler")
        self.worksheet.cell(row=last_row, column=1).font = Font(bold=True)
        
        self.worksheet.cell(row=last_row + 1, column=1, value="Taranan IP Bloğu:")
        self.worksheet.cell(row=last_row + 1, column=2, value=ip_block)
        
        self.worksheet.cell(row=last_row + 2, column=1, value="Bulunan Host Sayısı:")
        self.worksheet.cell(row=last_row + 2, column=2, value=total_hosts)
        
        self.worksheet.cell(row=last_row + 3, column=1, value="Tarama Tarihi:")
        self.worksheet.cell(row=last_row + 3, column=2, 
                           value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    
    def export(self, results: List[Dict], filename: str, ip_block: str = ""):
        """
        Sonuçları Excel dosyasına aktarır
        
        Args:
            results: Tarama sonuçları
            filename: Kaydedilecek dosya adı
            ip_block: Taranan IP bloğu
        """
        try:
            self.create_workbook()
            self.set_headers()
            self.add_data(results)
            
            if ip_block:
                self.add_summary(ip_block, len(results))
            
            self.workbook.save(filename)
            print(f"\n✓ Sonuçlar '{filename}' dosyasına kaydedildi.")
            return True
        except Exception as e:
            print(f"\n✗ Excel dosyası oluşturulurken hata: {e}")
            return False
